import os
import logging
import requests
import pandas as pd
from dotenv import load_dotenv
from azure.data.tables import TableServiceClient
from datetime import datetime

load_dotenv()
connection_string = os.getenv("storage_conn")
table_name = os.getenv("table_name")

# Predefined list of machines
ALLOWED_MACHINES = ['MC_PRESS_133', 'MC_PRESS_105']

# Define column order for better readability
fields = [
    'date',              # Date first
    'deviceType',        # Device information
    'shiftName',         # Shift information
    'startHour',         # Time information
    'endHour',
    'avg_oee',          # Performance metrics
    'avg_avail',
    'avg_perf',
    'avg_quality',
    'avg_current',
    'Ai_partcount',      # Production information
    'total_plannedPart',
    'avg_total_energy'   # Energy metrics
]

def get_available_machines(custID: str) -> list:
    """Fetch machine names for a customer"""
    try:
        url = "https://k2-oee-prod.azurewebsites.net/api/customer_devices?code=HOt2d_jYSqwXuML8FZ-OAK7UlrGnGsbTemJck3HhNHF1AzFuwOK3lQ=="
        payload = { "custID": custID }
        response = requests.post(url, json=payload)
        response.raise_for_status()
        data = response.json()
        machines = [d["device_name"] for d in data.get("devices", [])]
        return machines
    except Exception as e:
        logging.error(f" Failed to fetch machines: {e}")
        return []

def fetch_data_for_machine_and_dates(machine_name: str, start_date: str, end_date: str):
    """Query Azure Table for one machine between two dates and return data as a list of dictionaries"""
    try:
        service = TableServiceClient.from_connection_string(conn_str=connection_string)
        table_client = service.get_table_client(table_name=table_name)

        # Match machine and date range
        query_filter = f"device_name eq '{machine_name}' and date ge '{start_date}' and date le '{end_date}'"
        entities = table_client.query_entities(query_filter=query_filter)

        data_list = []
        for entity in entities:
            row_data = {'machine_name': machine_name}
            for field in fields:
                # Convert numeric values to float for calculations
                value = entity.get(field, 'N/A')
                if isinstance(value, (int, float)) or (isinstance(value, str) and value.replace('.', '').isdigit()):
                    try:
                        row_data[field] = float(value)
                    except ValueError:
                        row_data[field] = value
                else:
                    row_data[field] = value
            data_list.append(row_data)

        if not data_list:
            print(f" No records for machine '{machine_name}' between {start_date} and {end_date}")
        
        return data_list
    except Exception as e:
        print(f" Error fetching data for machine {machine_name}: {e}")
        return []

def create_daily_summary(df):
    """Create a daily summary showing only SH1 and SH2 totals for each day"""
    # Ensure date is datetime type
    df['date'] = pd.to_datetime(df['date'])
    
    # Ensure we only process SH1 and SH2
    df['shiftName'] = df['shiftName'].fillna('').str.upper()
    df = df[df['shiftName'].isin(['SH1', 'SH2'])]
    
    # Convert numeric strings to float where needed
    numeric_columns = ['avg_oee', 'avg_avail', 'avg_perf', 'avg_quality', 'avg_current', 
                      'Ai_partcount', 'total_plannedPart', 'avg_total_energy']
    
    for col in numeric_columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Sort by date and startHour to get correct start/end hours
    df = df.sort_values(['date', 'startHour'])
    
    # Create summary by date and shift
    summary_list = []
    
    for (date, shift), group in df.groupby(['date', 'shiftName']):
        summary = {
            'date': date,
            'shiftName': shift,
            'machine_name': ', '.join(sorted(set(group['machine_name']))),
            'deviceType': group['deviceType'].iloc[0],
            'startHour': group['startHour'].iloc[0],  # First startHour
            'endHour': group['endHour'].iloc[-1],     # Last endHour
            'avg_oee': group['avg_oee'].sum(),
            'avg_avail': group['avg_avail'].sum(),
            'avg_perf': group['avg_perf'].sum(),
            'avg_quality': group['avg_quality'].sum(),
            'avg_current': group['avg_current'].sum(),
            'Ai_partcount': group['Ai_partcount'].sum(),
            'total_plannedPart': group['total_plannedPart'].sum(),
            'avg_total_energy': group['avg_total_energy'].sum()
        }
        summary_list.append(summary)
    
    # Create DataFrame from summary list
    daily_summary = pd.DataFrame(summary_list)
    
    # Round numeric columns to 2 decimal places
    numeric_columns = ['avg_oee', 'avg_avail', 'avg_perf', 'avg_quality', 'avg_current', 
                      'Ai_partcount', 'total_plannedPart', 'avg_total_energy']
    daily_summary[numeric_columns] = daily_summary[numeric_columns].round(2)
    
    # Sort by date and shift
    daily_summary = daily_summary.sort_values(['date', 'shiftName'])
    
    # Reorder columns
    cols = ['date', 'shiftName', 'machine_name', 'deviceType', 'startHour', 'endHour'] + numeric_columns
    daily_summary = daily_summary[cols]
    
    return daily_summary

if __name__ == "__main__":
    custID = input("Enter Customer ID G46RK : ").strip()
    start_date = input("Enter Start Date (YYYY-MM-DD): ").strip()
    end_date = input("Enter End Date (YYYY-MM-DD): ").strip()

    print("\nAvailable machines:")
    for i, machine in enumerate(ALLOWED_MACHINES, 1):
        print(f"{i}. {machine}")
    
    # Get machine selection
    while True:
        try:
            selection = input("\nEnter machine numbers (comma-separated, e.g., 1,2) or 'all': ").strip().lower()
            if selection == 'all':
                selected_machines = ALLOWED_MACHINES
                break
            else:
                indices = [int(idx.strip()) - 1 for idx in selection.split(',')]
                selected_machines = [ALLOWED_MACHINES[i] for i in indices]
                if all(machine in ALLOWED_MACHINES for machine in selected_machines):
                    break
                else:
                    print("Invalid selection. Please try again.")
        except (ValueError, IndexError):
            print("Invalid input. Please enter valid numbers or 'all'.")

    print(f"\nFetching data for selected machines...")
    all_data = []
    
    # Fetch data for each selected machine
    for machine in selected_machines:
        machine_data = fetch_data_for_machine_and_dates(machine, start_date, end_date)
        all_data.extend(machine_data)
    
    if all_data:
        # Create DataFrame
        df = pd.DataFrame(all_data)
        
        # Reorder columns with machine_name first, then other fields in specified order
        column_order = ['machine_name'] + fields
        df = df.reindex(columns=column_order)
        
        # Create daily summary
        daily_summary = create_daily_summary(df.copy())
        
        # Create Excel writer with formatting
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = f"machine_data_{timestamp}.xlsx"
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Write detailed data
            df.to_excel(writer, index=False, sheet_name='Detailed Data')
            
            # Write daily summary
            daily_summary.to_excel(writer, index=False, sheet_name='Daily Summary')
            
            # Format both sheets
            workbook = writer.book
            from openpyxl.styles import Font, PatternFill, Alignment
            bold_font = Font(bold=True)
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            sh1_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')  # Light blue
            sh2_fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')  # Light green
            
            for sheet_name in ['Detailed Data', 'Daily Summary']:
                worksheet = writer.sheets[sheet_name]
                df_to_format = df if sheet_name == 'Detailed Data' else daily_summary
                
                # Format headers
                for col_idx, column in enumerate(df_to_format.columns, 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = bold_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Adjust column width to content
                    max_length = max(
                        df_to_format[column].astype(str).apply(len).max(),
                        len(str(column))
                    )
                    adjusted_width = min(max_length + 2, 50)  # Maximum width of 50
                    worksheet.column_dimensions[cell.column_letter].width = adjusted_width
                
                # Format shift rows in Daily Summary sheet
                if sheet_name == 'Daily Summary':
                    for row_idx in range(2, worksheet.max_row + 1):
                        shift = worksheet.cell(row=row_idx, column=2).value  # Column 2 is shiftName
                        fill = sh1_fill if shift == 'SH1' else sh2_fill
                        
                        # Center align and add background color to all cells in the row
                        for col_idx in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = Alignment(horizontal='center')
                            cell.fill = fill
        
        print(f"\nData exported successfully to {excel_file}")
        print("The Excel file contains two sheets:")
        print("1. 'Detailed Data' - All individual records")
        print("2. 'Daily Summary' - Daily aggregated values")
    else:
        print("\nNo data found for the selected machines in the specified date range.")
